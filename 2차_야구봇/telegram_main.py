# -*- coding: utf-8 -*-
"""
Created on Mon Sep 16 16:41:22 2019

@author: Onedas
"""

import json
import requests
from flask import Flask, request, Response
from openpyxl import load_workbook

# %%
from input_filter import dtp_filter
from input_filter import date_filter
import Schedule_crawler as SC
#import relay_crawler as RC
import ranking_crawler as RA
import get_relay_message as GM
#%%global variables
API_KEY = 'API_KEY'

HELP_TEAM_SEARCH = 'Team Search\n검색하고 싶은 날짜나 팀, 선수 이름을 입력해 주세요.'
HELP_RANKING_SEARCH = 'Ranking Search\n팀, 타자, 투수, 주요 랭킹이 있습니다.'


#%% DB
EXCEL_FILE_NAME = "DB.xlsx"
db = load_workbook(filename=EXCEL_FILE_NAME)
URL = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)
url_files = 'https://api.telegram.org/bot{token}/sendPhoto'.format(token=API_KEY)
user_db = db['User_DB']
keyboard = {                                        # Keyboard 형식
            'keyboard':[[{
                    'text': 'Team Search'
                        },
                    {'text': 'Ranking Search'
                        },
                    {'text': 'Clear'
                        }]
                    ],
            'one_time_keyboard' : False
            }

TEAMLIST= []
PLAYERLIST=[]

# %%
app = Flask(__name__)
    
def parse_message(data):
    '''응답data 로부터 chat_id 와 text, user_name을 추출.'''
    
    chat_id = data['message']['chat']['id']
    msg = data['message']['text']
    user_name = data['message']['chat']['first_name'] + data['message']['chat']['last_name']
    
    return chat_id, msg, user_name    #https://core.telegram.org/bots/api#keyboardbutton

def find_user_row(user_id, user_name):
    ''' DB에서 user_id 로 부터 DB번호를 반환
        없으면 DB에 업데이트
    '''
    user_exist = False
    
    for row in user_db.rows:
        if row[0].value == user_id:
            user_exist = True
            user_row = row[0].row
            break #return user_row
    
    if not user_exist:
        user_db[user_db.max_row + 1][0].value = user_id
        
        user_db[user_db.max_row][1].value = user_name
        user_db[user_db.max_row][2].value = 1
        user_db[user_db.max_row][3].value = 'reset' #button
        user_db[user_db.max_row][4].value = None #date
        user_db[user_db.max_row][5].value = None #team
        user_db[user_db.max_row][6].value = None #player
        user_row = user_db.max_row
        db.save(EXCEL_FILE_NAME)    

    return user_row

def update_data(msg, user_row):
    '''DB와 msg를 비교하여 date, team, player 선택'''    
    prev_date = user_db[user_row][4].value
    prev_team = user_db[user_row][5].value
    prev_player = user_db[user_row][6].value

    date, team, player,dtp_state = dtp_filter(msg)

    if date == None:
        date = prev_date
    
    if team == None:
        team = prev_team
        
    if player == None:
        player = prev_player
    
    user_db[user_row][4].value = date
    user_db[user_row][5].value = team
    user_db[user_row][6].value = player
    db.save(EXCEL_FILE_NAME)
    
    return date, team, player, dtp_state


def is_date_team_player(date,team,player):
    # 날짜와 팀, 선수 이름이 있는지 체크
    state=0
    if date!=None:
        state +=1
        
    if team!=None:
        state +=2
    
    if player != None:
        state +=4
    
    return state

def reset_DB(user_row):
    user_db[user_row][3].value='reset'
    user_db[user_row][4].value=None
    user_db[user_row][5].value=None
    user_db[user_row][6].value=None
    db.save(EXCEL_FILE_NAME)


################################################################################

def button_message(chat_id,user_row, text=''):
    """
    Chat-id 와 text를 변수로 받아 메세지를 보내주는데
    params 안에 키보드를 설정해서 같이 보내주는 방법
    """
    url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)   #sendMessage
    
    params = {'chat_id':chat_id, 'text':'', 'reply_markup' : keyboard}
    requests.post(url, json=params)
    prev_button = user_db[user_row][3].value
    
    
    if text not in ['Team Search', 'Ranking Search','Clear'] and prev_button == 'reset':
        params = {'chat_id':chat_id, 'text': '안녕하세요. KBOT입니다.\n버튼을 눌러주세요', 'reply_markup' : keyboard}
        requests.post(url, json=params)
    
    if text=='Team Search':
        params = {'chat_id':chat_id, 'text': HELP_TEAM_SEARCH, 'reply_markup' : keyboard}
        user_db[user_row][3].value = text
        db.save(EXCEL_FILE_NAME)    
        requests.post(url, json=params)
    
    if text=='Ranking Search':
        params = {'chat_id':chat_id, 'text':HELP_RANKING_SEARCH, 'reply_markup' : keyboard}
        requests.post(url, json=params)
        user_db[user_row][3].value = text
        db.save(EXCEL_FILE_NAME)    
    
    if text =='Clear':
        params = {'chat_id':chat_id, 'text':'초기화 합니다', 'reply_markup' : keyboard}
        requests.post(url, json=params)
        user_db[user_row][3].value = text
        db.save(EXCEL_FILE_NAME)            
    
    return user_db[user_row][3].value

##################################################################################################
def Team_search_senario(msg,chat_id,dtp_state, date, team, player):
#    dtp_state = is_date_team_player(date,team,player)
    text=''
    if dtp_state ==0: # 아무것도 없는 경우
        text += '날짜, 팀, 선수이름을 입력해주세요.'
        
    elif dtp_state ==1: #날짜
        text +='{}의 경기를 모두 출력합니다\n'.format(date)
        text += SC.print_dailySchedule(date)
        params = {'chat_id':chat_id, 'text': text}#, 'reply_markup' : keyboard}
        requests.post(URL, json=params)

    elif dtp_state ==2: #팀
        text += '해당 {}의 경기를 모두 출력합니다\n'.format(team)
        params = {'chat_id':chat_id, 'text': text}#, 'reply_markup' : keyboard}
        requests.post(URL, json=params)
    
    elif dtp_state ==3: #날짜, 팀
        text += '해당 {}의 {}의 경기결과를 출력합니다\n'.format(date,team)
        url = SC.get_game_url(date,team)

        text += GM.get_team_message(date,team)
#        text += RC.print_teamHistory(team,url)[:-100]
        
        params = {'chat_id':chat_id, 'text': text}#, 'reply_markup' : keyboard}
        requests.post(URL, json=params)
        path = GM.get_game_graph(date,team)
        files = {'photo': open(path, 'rb')}
        data = {'chat_id': chat_id}
        requests.post(url_files, files=files, data=data)


    
    elif dtp_state ==4: #타자
        text += '{} 타자 정보만 있습니다\n'.format(player)
        params = {'chat_id':chat_id, 'text': text}#, 'reply_markup' : keyboard}
        requests.post(URL, json=params)
    
    elif dtp_state ==5: #날짜, 타자
        text += '{}와 {} 정보만 있습니다.\n'.format(date,player)
        params = {'chat_id':chat_id, 'text': text}#, 'reply_markup' : keyboard}
        requests.post(URL, json=params)
    
    elif dtp_state ==6: # 팀, 타자
        text += '{}와 {} 정보만 있습니다.\n'.format(team,player)
        params = {'chat_id':chat_id, 'text': text}#, 'reply_markup' : keyboard}
        requests.post(URL, json=params)
    
    else:# dtp_state ==7: #팀, 타자, 날짜
        text +='해당 {}와 {}의 {} 타자의 경기 결과를 출력합니다.\n'.format(date,team,player)

        text+=GM.get_player_message(date,team,player)
        params = {'chat_id':chat_id, 'text': text}#, 'reply_markup' : keyboard}
        requests.post(URL, json=params)
    
    


def Ranking_search_senario(msg,chat_id):
    if '주요' in msg:
        text = '팀랭킹을 출력합니다.\n'
        text += RA.load_Ranking(keys=['Main'])
        params = {'chat_id':chat_id, 'text': text}#, 'reply_markup' : keyboard}
        requests.post(URL, json=params)

    if '팀' in msg:
        text = '팀랭킹을 출력합니다.\n'
        text += RA.load_Ranking(keys=['Team'])
        params = {'chat_id':chat_id, 'text': text}#, 'reply_markup' : keyboard}
        requests.post(URL, json=params)

    if '타자' in msg:
        text = '타자랭킹을 출력합니다.\n'
        text += RA.load_Ranking(keys=['Batter'])
        params = {'chat_id':chat_id, 'text': text}#, 'reply_markup' : keyboard}
        requests.post(URL, json=params)

    if '투수' in msg:
        text = '투수랭킹을 출력합니다.\n'
        text += RA.load_Ranking(keys=['Pitcher'])
        params = {'chat_id':chat_id, 'text': text}#, 'reply_markup' : keyboard}
        requests.post(URL, json=params)


##################################################################################################
# 경로 설정, URL 설정
@app.route('/', methods=['POST', 'GET'])
def main():
    if request.method == 'POST':

        message = request.get_json()
        chat_id, msg, chat_name = parse_message(message)
        user_row = find_user_row(chat_id, chat_name)
        
#        prev_button, prev_date, prev_team, prev_player = find_prev_data(user_row)
        
        
        date, team, player,keyword_state = update_data(msg,user_row)
        dtp_state = is_date_team_player(date,team,player) # 0 ~ 7 까지

        button = button_message(chat_id,user_row, msg)
        
        print(button,date,team,player)
        print(dtp_state)
        if button == None:
            return 0
        
        if button == 'Team Search':
            if keyword_state:
                Team_search_senario(msg, chat_id,dtp_state, date,team, player)
            else:
                params = {'chat_id': chat_id, 'text': 'ex)9월1일 sk 경기 알려줘'}  # , 'reply_markup' : keyboard}
                requests.post(URL, json=params)

        if button == 'Ranking Search':
            Ranking_search_senario(msg,chat_id)
        
        if button == 'Clear':
            reset_DB(user_row)
        
        return Response('ok', status=200)

    else:
        return 'Hello World!'


# Python 에서는 실행시킬때 __name__ 이라는 변수에
# __main__ 이라는 값이 할당
if __name__ == '__main__':
    app.run(port = 5000)
