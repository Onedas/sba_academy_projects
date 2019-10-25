# -*- coding: utf-8 -*-
"""
Created on Wed Sep 18 11:07:33 2019

@author: Onedas
"""

from openpyxl import load_workbook

EXCEL_FILE_NAME = 'DB.xlsx'
db = load_workbook(filename=EXCEL_FILE_NAME)
#tuto_db = db['Tuto']

# create_sheet 로 새로운 Sheet 생성 가능
# tuto_db 는 Sheet 객체이므로, 엑셀 객체인 db를 가지고 사용
# .title 로 Sheet의 이름을 변경 가능

db.remove_sheet(db['User_DB'])
#
create_db = db.create_sheet('Create_sheet')
create_db.title = 'User_DB'

user_db = db['User_DB']
user_db['A'][0].value= 'ID'
user_db['B'][0].value= 'NAME'
user_db['C'][0].value= 'COUNT'
user_db['D'][0].value= 'BUTTON'
user_db['E'][0].value= 'DATE'
user_db['F'][0].value= 'TEAM'
user_db['G'][0].value= 'PLAYER'


db.save(EXCEL_FILE_NAME)
print('ok')